"""Structured data import: preview, validation, duplicate detection, commit.

The application imports CSV / XLSX exports prepared from the Bassignana source
documents. It deliberately does NOT read PDFs: complex programme PDFs are
registered as evidence, never machine-interpreted.

Every import is a two-step operation. Nothing is written to the database until
the user confirms a validated preview.
"""
from __future__ import annotations

import csv
import io
import os
import re
import uuid
from datetime import date, datetime
from pathlib import Path

from flask import current_app

from app import constants as C
from app.extensions import db
from app.i18n import translate as t
from app.models import (
    AcceptanceGate,
    AcceptanceGateItem,
    ActivityQuantity,
    Area,
    DocumentRegisterItem,
    ImportBatch,
    InspectionRequirement,
    Material,
    PaymentMilestone,
    PermitItem,
    ProcurementPackage,
    ScheduleVersion,
    SiteObservation,
    SourceDocument,
    WbsActivity,
)

TRUE_TOKENS = {"1", "y", "yes", "true", "t", "si", "sì", "x", "evet"}
FALSE_TOKENS = {"0", "n", "no", "false", "f", ""}

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y",
    "%d/%m/%y", "%d.%m.%y", "%Y/%m/%d", "%m/%d/%Y",
]


# --------------------------------------------------------------------------
# Value coercion
# --------------------------------------------------------------------------
def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def to_str(value, max_length=None):
    text = clean(value)
    if not isinstance(text, str):
        text = str(text)
    text = text.strip()
    if max_length and len(text) > max_length:
        text = text[:max_length]
    return text


def to_float(value, field=None, errors=None):
    text = clean(value)
    if text in ("", None):
        return None
    if isinstance(text, (int, float)):
        return float(text)
    text = str(text).replace("%", "").strip()
    # Accept both 1.234,56 and 1,234.56
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in ("", "-", "."):
        if errors is not None:
            errors.append(f"{field or 'value'}: '{clean(value)}' is not a number")
        return None
    try:
        return float(text)
    except ValueError:
        if errors is not None:
            errors.append(f"{field or 'value'}: '{clean(value)}' is not a number")
        return None


def to_int(value, field=None, errors=None):
    result = to_float(value, field, errors)
    return None if result is None else int(round(result))


def to_date(value, field=None, errors=None):
    text = clean(value)
    if text in ("", None):
        return None
    if isinstance(text, datetime):
        return text.date()
    if isinstance(text, date):
        return text
    text = str(text).strip()
    # Strip an Italian weekday prefix such as "lun 11/05/26".
    text = re.sub(r"^[a-zA-Zàèéìòù]{3}\s+", "", text)
    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt).date()
            if fmt.endswith("%y") and parsed.year < 1980:
                parsed = parsed.replace(year=parsed.year + 100)
            return parsed
        except ValueError:
            continue
    if errors is not None:
        errors.append(f"{field or 'date'}: '{text}' is not a recognised date "
                      f"(use YYYY-MM-DD or DD/MM/YYYY)")
    return None


def to_bool(value, default=False):
    text = clean(value)
    if text in ("", None):
        return default
    if isinstance(text, bool):
        return text
    return str(text).strip().lower() in TRUE_TOKENS


def normalise_header(header):
    return re.sub(r"[^a-z0-9]+", "_", str(header or "").strip().lower()).strip("_")


# --------------------------------------------------------------------------
# File handling
# --------------------------------------------------------------------------
#: Uploaded import files are kept only long enough to preview and commit them.
IMPORT_RETENTION_HOURS = 24


def import_workspace():
    path = Path(current_app.config["DATA_DIR"]) / "imports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def purge_stale_uploads(max_age_hours=None):
    """Delete import uploads older than the retention window.

    An import that is previewed but never confirmed would otherwise leave its
    file behind forever. Nothing here touches committed data.
    """
    import time

    max_age = (IMPORT_RETENTION_HOURS if max_age_hours is None else max_age_hours) * 3600
    now = time.time()
    removed = 0
    try:
        entries = list(import_workspace().iterdir())
    except OSError:  # pragma: no cover - defensive
        return 0
    for path in entries:
        try:
            if path.is_file() and (now - path.stat().st_mtime) >= max_age:
                path.unlink()
                removed += 1
        except OSError:  # pragma: no cover - file in use
            continue
    return removed


def save_upload(file_storage):
    """Persist an uploaded import file and return its token."""
    suffix = Path(file_storage.filename or "import.csv").suffix.lower()
    if suffix not in current_app.config["ALLOWED_IMPORT_EXT"]:
        raise ValueError(t("Unsupported import file type '{suffix}'. Use .csv or .xlsx.",
                           suffix=suffix))
    token = uuid.uuid4().hex
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", Path(file_storage.filename or "import").name)
    target = import_workspace() / f"{token}__{safe_name}"
    file_storage.save(target)
    return token, target.name


def resolve_upload(token):
    for path in import_workspace().iterdir():
        if path.name.startswith(f"{token}__"):
            return path
    raise FileNotFoundError(
        t("Uploaded import file is no longer available. Please upload it again."))


def read_rows(path):
    """Read a CSV or XLSX file into (headers, list-of-dicts)."""
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(path):
    raw = Path(path).read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [normalise_header(h) for h in (reader.fieldnames or [])]
    rows = []
    for raw_row in reader:
        rows.append({normalise_header(k): v for k, v in raw_row.items() if k is not None})
    return headers, rows


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl ships in requirements
        raise RuntimeError(t("openpyxl is required to read .xlsx import files")) from exc
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []
    headers = [normalise_header(h) for h in header_row]
    rows = []
    for values in iterator:
        if values is None or all(v in (None, "") for v in values):
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    workbook.close()
    return headers, rows


# --------------------------------------------------------------------------
# Import type definitions
# --------------------------------------------------------------------------
class ImportSpec:
    """Declarative description of one import template."""

    def __init__(self, key, label, columns, required, description, guidance=""):
        self.key = key
        self.label = label
        self.columns = columns
        self.required = required
        self.description = description
        self.guidance = guidance

    def template_csv(self):
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow(self.columns)
        return buffer.getvalue()


SPECS = {
    "schedule": ImportSpec(
        key="schedule",
        label="Schedule / WBS",
        columns=[
            "wbs_code", "parent_wbs_code", "activity_name", "work_package", "discipline",
            "contractual_start", "contractual_finish", "current_planned_start",
            "current_planned_finish", "duration_days", "milestone", "unit",
            "total_required_quantity", "progress_weight", "current_reported_completion_pct",
            "responsible_party", "schedule_version", "schedule_status", "source_document",
        ],
        required=["wbs_code", "activity_name"],
        description="One row per activity of a single programme revision.",
        guidance=(
            "Dates accept YYYY-MM-DD or DD/MM/YYYY. Contractual dates are written only when the "
            "target schedule version is a CONTRACTUAL BASELINE; importing a working programme "
            "never overwrites contractual dates."
        ),
    ),
    "areas": ImportSpec(
        key="areas",
        label="Areas / Workfronts",
        columns=["area_code", "area_name", "description", "parent_area", "drawing_reference",
                 "ifc_revision", "active"],
        required=["area_code", "area_name"],
        description="Workfront register taken from the approved IFC / general layout.",
        guidance="Area codes must come from approved project layout data. Nothing is invented.",
    ),
    "quantities": ImportSpec(
        key="quantities",
        label="Quantities / BOQ",
        columns=["wbs_code", "area_code", "activity_name", "item", "total_quantity", "unit",
                 "source", "revision", "notes"],
        required=["item", "total_quantity"],
        description="Approved quantity register. Quantities are never inferred.",
    ),
    "materials": ImportSpec(
        key="materials",
        label="Materials / Equipment Packages",
        columns=["package_code", "package_name", "item", "manufacturer", "vendor",
                 "approved_vendor", "unit", "contract_quantity", "required_quantity",
                 "ordered_quantity", "planned_delivery", "actual_delivery", "po_reference",
                 "fat_required", "source"],
        required=["package_code", "package_name"],
        description="Procurement packages and their material lines.",
        guidance="Delivered, accepted and installed quantities are recorded on site, "
                 "not imported as one another.",
    ),
    "quality": ImportSpec(
        key="quality",
        label="Quality Requirements / ITP",
        columns=["wbs_code", "work_package", "itp_reference", "inspection_type", "point_type",
                 "required_evidence", "acceptance_criterion", "applicable_specification",
                 "discipline", "source"],
        required=["itp_reference"],
        description="ITP / checklist points, hold, witness and review points.",
    ),
    "acceptance": ImportSpec(
        key="acceptance",
        label="Acceptance Gate Items",
        columns=["gate_code", "item_code", "item_name", "description", "category",
                 "responsible_party", "target_date", "contract_reference", "source"],
        required=["gate_code", "item_name"],
        description="Prerequisites of Gates A-D from Schedules 04A / 04B / 04C / 04D.",
    ),
    "permits": ImportSpec(
        key="permits",
        label="Permits / Readiness",
        columns=["item_name", "authority", "responsibility", "required_for", "required_by_date",
                 "issued_date", "expiry_date", "status", "document_reference", "blocker_impact",
                 "comments", "source"],
        required=["item_name"],
        description="Permit and readiness register. Status register only.",
        guidance="Import only a verified Bassignana permit register.",
    ),
    "payments": ImportSpec(
        key="payments",
        label="Payment Milestones",
        columns=["sequence", "milestone_code", "description", "percentage", "wbs_code",
                 "gate_code", "package_code", "planned_date", "forecast_date", "status",
                 "comments", "source"],
        required=["description", "percentage"],
        description="Contractual payment milestones from Schedule 10.",
        guidance="Percentages must total 100%. Amounts are derived from the Contract "
                 "Price registered in Project Setup and are never typed in.",
    ),
    "observations": ImportSpec(
        key="observations",
        label="Site Observations",
        columns=["observation_date", "area_code", "wbs_code", "observation", "category",
                 "severity", "action_required", "responsible_party", "target_date",
                 "status", "source"],
        required=["observation_date", "observation"],
        description="Backlog of site observations from existing site reports.",
        guidance="Use this to load observations recorded before the system went live. "
                 "New observations are recorded in the Daily Site Diary.",
    ),
    "documents": ImportSpec(
        key="documents",
        label="Document Register",
        columns=["document_number", "title", "category", "discipline", "wbs_code", "revision",
                 "status", "required_date", "gate_code", "mandatory", "folder_path", "remarks"],
        required=["title"],
        description="Required project evidence documents, optionally tied to an acceptance gate.",
    ),
}


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------
def validate(import_key, rows, options=None):
    """Validate parsed rows and classify each as CREATE, UPDATE, SKIP or ERROR."""
    options = options or {}
    spec = SPECS[import_key]
    handler = _VALIDATORS[import_key]

    results = []
    seen_keys = {}
    for index, raw in enumerate(rows, start=2):  # row 1 is the header
        errors, warnings = [], []
        missing = [c for c in spec.required if not to_str(raw.get(c))]
        if missing:
            errors.append("missing required value(s): " + ", ".join(missing))
            results.append({"index": index, "raw": raw, "data": {}, "errors": errors,
                            "warnings": warnings, "action": "ERROR", "key": None})
            continue

        data, key, action, row_warnings, row_errors = handler(raw, options)
        warnings.extend(row_warnings)
        errors.extend(row_errors)

        if key is not None:
            if key in seen_keys:
                errors.append(f"duplicate of row {seen_keys[key]} in this file (key '{key}')")
                action = "ERROR"
            else:
                seen_keys[key] = index

        results.append({
            "index": index, "raw": raw, "data": data, "errors": errors,
            "warnings": warnings,
            "action": "ERROR" if errors else action,
            "key": key,
        })

    summary = {
        "total": len(results),
        "create": sum(1 for r in results if r["action"] == "CREATE"),
        "update": sum(1 for r in results if r["action"] == "UPDATE"),
        "skip": sum(1 for r in results if r["action"] == "SKIP"),
        "error": sum(1 for r in results if r["action"] == "ERROR"),
        "warning": sum(1 for r in results if r["warnings"]),
    }
    return {"spec": spec, "rows": results, "summary": summary}


def _v_schedule(raw, options):
    errors, warnings = [], []
    version = options.get("version")
    wbs = to_str(raw.get("wbs_code"), 60)
    data = {
        "wbs_code": wbs,
        "parent_wbs_code": to_str(raw.get("parent_wbs_code"), 60) or _derive_parent(wbs),
        "activity_name": to_str(raw.get("activity_name"), 400),
        "work_package": to_str(raw.get("work_package"), 200),
        "discipline": to_str(raw.get("discipline"), 60),
        "contractual_start": to_date(raw.get("contractual_start"), "contractual_start", errors),
        "contractual_finish": to_date(raw.get("contractual_finish"), "contractual_finish", errors),
        "plan_start": to_date(raw.get("current_planned_start"), "current_planned_start", errors),
        "plan_finish": to_date(raw.get("current_planned_finish"), "current_planned_finish", errors),
        "duration_days": to_float(raw.get("duration_days"), "duration_days", errors),
        "is_milestone": to_bool(raw.get("milestone")),
        "unit": to_str(raw.get("unit"), 30),
        "total_required_quantity": to_float(raw.get("total_required_quantity"),
                                            "total_required_quantity", errors),
        "progress_weight": to_float(raw.get("progress_weight"), "progress_weight", errors),
        "reported_completion_pct": to_float(raw.get("current_reported_completion_pct"),
                                            "current_reported_completion_pct", errors),
        "responsible_party": to_str(raw.get("responsible_party"), 160),
        "source_reference": to_str(raw.get("source_document"), 300),
    }
    if data["duration_days"] == 0:
        data["is_milestone"] = True
    if data["reported_completion_pct"] is not None:
        if not (0 <= data["reported_completion_pct"] <= 100):
            errors.append("current_reported_completion_pct must be between 0 and 100")
    if data["plan_start"] and data["plan_finish"] and data["plan_finish"] < data["plan_start"]:
        errors.append("current planned finish is before current planned start")
    if data["contractual_start"] and data["contractual_finish"] \
            and data["contractual_finish"] < data["contractual_start"]:
        errors.append("contractual finish is before contractual start")
    if not data["plan_start"] and not data["contractual_start"]:
        warnings.append("no start date supplied")

    action = "CREATE"
    if version is not None:
        existing = WbsActivity.query.filter_by(
            schedule_version_id=version.id, wbs_code=wbs).first()
        if existing is not None:
            action = "UPDATE"
            if version.locked:
                errors.append("schedule version is locked; unlock it before re-importing")
    return data, wbs, action, warnings, errors


def _derive_parent(wbs_code):
    if not wbs_code or "." not in wbs_code:
        return ""
    return wbs_code.rsplit(".", 1)[0]


def _v_areas(raw, options):
    errors, warnings = [], []
    code = to_str(raw.get("area_code"), 60)
    data = {
        "area_code": code,
        "area_name": to_str(raw.get("area_name"), 200),
        "description": to_str(raw.get("description")),
        "parent_area": to_str(raw.get("parent_area"), 60),
        "drawing_reference": to_str(raw.get("drawing_reference"), 200),
        "ifc_revision": to_str(raw.get("ifc_revision"), 60),
        "active": to_bool(raw.get("active"), default=True),
    }
    if not data["drawing_reference"]:
        warnings.append("no drawing reference: the workfront will not be traceable to an IFC document")
    existing = Area.query.filter_by(area_code=code).first()
    return data, code, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_quantities(raw, options):
    errors, warnings = [], []
    item = to_str(raw.get("item"), 300)
    wbs = to_str(raw.get("wbs_code"), 60)
    area_code = to_str(raw.get("area_code"), 60)
    quantity = to_float(raw.get("total_quantity"), "total_quantity", errors)
    if quantity is not None and quantity < 0:
        errors.append("total_quantity cannot be negative")
    area = Area.query.filter_by(area_code=area_code).first() if area_code else None
    if area_code and area is None:
        errors.append(f"area '{area_code}' is not registered; import the area register first")
    data = {
        "wbs_code": wbs,
        "area_id": area.id if area else None,
        "activity_name": to_str(raw.get("activity_name"), 300),
        "item": item,
        "total_quantity": quantity,
        "unit": to_str(raw.get("unit"), 30),
        "source_reference": to_str(raw.get("source"), 300),
        "revision": to_str(raw.get("revision"), 60),
        "notes": to_str(raw.get("notes")),
    }
    key = f"{wbs}|{area_code}|{item}"
    existing = ActivityQuantity.query.filter_by(wbs_code=wbs or None, item=item).first()
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_materials(raw, options):
    errors, warnings = [], []
    package_code = to_str(raw.get("package_code"), 40)
    item = to_str(raw.get("item"), 300)
    data = {
        "package_code": package_code,
        "package_name": to_str(raw.get("package_name"), 200),
        "item": item,
        "manufacturer": to_str(raw.get("manufacturer"), 200),
        "vendor": to_str(raw.get("vendor"), 200),
        "approved_vendor": to_bool(raw.get("approved_vendor")),
        "unit": to_str(raw.get("unit"), 30),
        "contract_quantity": to_float(raw.get("contract_quantity"), "contract_quantity", errors),
        "total_required": to_float(raw.get("required_quantity"), "required_quantity", errors),
        "ordered": to_float(raw.get("ordered_quantity"), "ordered_quantity", errors),
        "planned_delivery": to_date(raw.get("planned_delivery"), "planned_delivery", errors),
        "actual_delivery": to_date(raw.get("actual_delivery"), "actual_delivery", errors),
        "po_reference": to_str(raw.get("po_reference"), 160),
        "fat_required": to_bool(raw.get("fat_required")),
        "source_reference": to_str(raw.get("source"), 300),
    }
    if item and data["total_required"] is None:
        warnings.append(f"{C.DATA_REQUIRED}: required quantity for this material line")
    key = f"{package_code}|{item}"
    if item:
        existing = (Material.query.join(ProcurementPackage)
                    .filter(ProcurementPackage.package_code == package_code,
                            Material.item == item).first())
    else:
        existing = ProcurementPackage.query.filter_by(package_code=package_code).first()
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_quality(raw, options):
    errors, warnings = [], []
    itp = to_str(raw.get("itp_reference"), 120)
    wbs = to_str(raw.get("wbs_code"), 60)
    point_type = (to_str(raw.get("point_type")) or "REVIEW").upper()
    if point_type not in C.INSPECTION_POINT_TYPES:
        warnings.append(f"point_type '{point_type}' is not one of "
                        f"{', '.join(C.INSPECTION_POINT_TYPES)}; stored as REVIEW")
        point_type = "REVIEW"
    data = {
        "wbs_code": wbs,
        "work_package": to_str(raw.get("work_package"), 200),
        "itp_reference": itp,
        "inspection_type": to_str(raw.get("inspection_type"), 80),
        "point_type": point_type,
        "required_evidence": to_str(raw.get("required_evidence")),
        "acceptance_criterion": to_str(raw.get("acceptance_criterion")),
        "applicable_specification": to_str(raw.get("applicable_specification"), 300),
        "discipline": to_str(raw.get("discipline"), 60),
        "source_reference": to_str(raw.get("source"), 300),
    }
    key = f"{itp}|{wbs}|{data['inspection_type']}"
    existing = InspectionRequirement.query.filter_by(
        itp_reference=itp, wbs_code=wbs or None,
        inspection_type=data["inspection_type"] or None).first()
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_acceptance(raw, options):
    errors, warnings = [], []
    gate_code = to_str(raw.get("gate_code"), 4).upper()
    item_name = to_str(raw.get("item_name"), 400)
    if gate_code not in {g[0] for g in C.GATES}:
        errors.append(f"gate_code '{gate_code}' must be one of A, B, C, D")
    data = {
        "gate_code": gate_code,
        "item_code": to_str(raw.get("item_code"), 40),
        "item_name": item_name,
        "description": to_str(raw.get("description")),
        "category": to_str(raw.get("category"), 80),
        "responsible_party": to_str(raw.get("responsible_party"), 160),
        "target_date": to_date(raw.get("target_date"), "target_date", errors),
        "contract_reference": to_str(raw.get("contract_reference"), 200),
        "source_reference": to_str(raw.get("source"), 300),
    }
    key = f"{gate_code}|{item_name}"
    gate = AcceptanceGate.query.filter_by(gate_code=gate_code).first()
    existing = (AcceptanceGateItem.query.filter_by(gate_id=gate.id, item_name=item_name).first()
                if gate else None)
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_permits(raw, options):
    errors, warnings = [], []
    name = to_str(raw.get("item_name"), 300)
    status = (to_str(raw.get("status")) or "NOT STARTED").upper()
    if status not in C.PERMIT_STATUS:
        warnings.append(f"status '{status}' is not a permit status; stored as NOT STARTED")
        status = "NOT STARTED"
    data = {
        "item_name": name,
        "authority": to_str(raw.get("authority"), 200),
        "responsibility": to_str(raw.get("responsibility"), 160),
        "required_for": to_str(raw.get("required_for"), 300),
        "required_by_date": to_date(raw.get("required_by_date"), "required_by_date", errors),
        "issued_date": to_date(raw.get("issued_date"), "issued_date", errors),
        "expiry_date": to_date(raw.get("expiry_date"), "expiry_date", errors),
        "status": status,
        "document_reference": to_str(raw.get("document_reference"), 300),
        "blocker_impact": to_str(raw.get("blocker_impact")),
        "comments": to_str(raw.get("comments")),
        "source_reference": to_str(raw.get("source"), 300),
    }
    existing = PermitItem.query.filter_by(item_name=name).first()
    return data, name, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_payments(raw, options):
    errors, warnings = [], []
    description = to_str(raw.get("description"), 400)
    code = to_str(raw.get("milestone_code"), 40)
    percentage = to_float(raw.get("percentage"), "percentage", errors)
    if percentage is not None and not (0 <= percentage <= 100):
        errors.append("percentage must be between 0 and 100")
    status = (to_str(raw.get("status")) or "NOT STARTED").upper()
    from app.services.commercial_service import PAYMENT_STATUS
    if status not in PAYMENT_STATUS:
        warnings.append(f"status '{status}' is not a payment status; stored as NOT STARTED")
        status = "NOT STARTED"
    gate_code = to_str(raw.get("gate_code"), 4).upper() or None
    if gate_code and gate_code not in {g[0] for g in C.GATES}:
        warnings.append(f"gate_code '{gate_code}' is not A/B/C/D; stored without a gate link")
        gate_code = None
    data = {
        "sequence": to_int(raw.get("sequence")) or 0,
        "milestone_code": code,
        "description": description,
        "percentage": percentage or 0.0,
        "wbs_code": to_str(raw.get("wbs_code"), 60),
        "gate_code": gate_code,
        "package_code": to_str(raw.get("package_code"), 40),
        "planned_date": to_date(raw.get("planned_date"), "planned_date", errors),
        "forecast_date": to_date(raw.get("forecast_date"), "forecast_date", errors),
        "status": status,
        "comments": to_str(raw.get("comments")),
        "source_reference": to_str(raw.get("source"), 300),
    }
    key = code or description
    existing = (PaymentMilestone.query.filter_by(milestone_code=code).first() if code
                else PaymentMilestone.query.filter_by(description=description).first())
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_observations(raw, options):
    errors, warnings = [], []
    observation_date = to_date(raw.get("observation_date"), "observation_date", errors)
    observation = to_str(raw.get("observation"))
    area_code = to_str(raw.get("area_code"), 60)
    area = Area.query.filter_by(area_code=area_code).first() if area_code else None
    if area_code and area is None:
        errors.append(f"area '{area_code}' is not registered; import the area register first")
    category = to_str(raw.get("category")) or "Other"
    if category not in C.OBSERVATION_CATEGORIES:
        warnings.append(f"category '{category}' is not a recognised observation category; "
                        f"stored as Other")
        category = "Other"
    severity = (to_str(raw.get("severity")) or "LOW").upper()
    if severity not in C.SEVERITIES:
        warnings.append(f"severity '{severity}' is not LOW/MEDIUM/HIGH/CRITICAL; stored as LOW")
        severity = "LOW"
    status = (to_str(raw.get("status")) or "OPEN").upper()
    if status not in C.ACTION_STATUS:
        warnings.append(f"status '{status}' is not an action status; stored as OPEN")
        status = "OPEN"
    data = {
        "entry_date": observation_date,
        "area_id": area.id if area else None,
        "wbs_code": to_str(raw.get("wbs_code"), 60),
        "observation": observation,
        "category": category,
        "severity": severity,
        "action_required": to_str(raw.get("action_required")),
        "responsible_party": to_str(raw.get("responsible_party"), 160),
        "target_date": to_date(raw.get("target_date"), "target_date", errors),
        "status": status,
        "source_reference": to_str(raw.get("source"), 300),
    }
    key = f"{observation_date}|{observation[:80]}"
    existing = SiteObservation.query.filter_by(
        entry_date=observation_date, observation=observation).first()
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


def _v_documents(raw, options):
    errors, warnings = [], []
    title = to_str(raw.get("title"), 400)
    number = to_str(raw.get("document_number"), 120)
    gate_code = to_str(raw.get("gate_code"), 4).upper() or None
    if gate_code and gate_code not in {g[0] for g in C.GATES}:
        warnings.append(f"gate_code '{gate_code}' is not A/B/C/D; stored without a gate link")
        gate_code = None
    status = (to_str(raw.get("status")) or "NOT STARTED").upper()
    if status not in C.DOCUMENT_STATUS:
        warnings.append(f"status '{status}' is not a document status; stored as NOT STARTED")
        status = "NOT STARTED"
    data = {
        "document_number": number,
        "title": title,
        "category": to_str(raw.get("category"), 60),
        "discipline": to_str(raw.get("discipline"), 60),
        "wbs_code": to_str(raw.get("wbs_code"), 60),
        "revision": to_str(raw.get("revision"), 60),
        "status": status,
        "required_date": to_date(raw.get("required_date"), "required_date", errors),
        "gate_code": gate_code,
        "mandatory": to_bool(raw.get("mandatory"), default=True),
        "folder_path": to_str(raw.get("folder_path"), 400),
        "remarks": to_str(raw.get("remarks")),
    }
    key = f"{number}|{title}|{gate_code}"
    existing = DocumentRegisterItem.query.filter_by(
        title=title, gate_code=gate_code, document_number=number or None).first()
    return data, key, ("UPDATE" if existing else "CREATE"), warnings, errors


_VALIDATORS = {
    "schedule": _v_schedule,
    "areas": _v_areas,
    "quantities": _v_quantities,
    "materials": _v_materials,
    "quality": _v_quality,
    "acceptance": _v_acceptance,
    "permits": _v_permits,
    "documents": _v_documents,
    "payments": _v_payments,
    "observations": _v_observations,
}


# --------------------------------------------------------------------------
# Commit
# --------------------------------------------------------------------------
def commit(import_key, validation, options=None, filename=None):
    """Write validated rows. Rows classified ERROR are always skipped."""
    options = options or {}
    source_document = options.get("source_document")
    version = options.get("version")
    committer = _COMMITTERS[import_key]

    batch = ImportBatch(
        import_type=import_key,
        filename=filename,
        source_document_id=source_document.id if source_document else None,
        schedule_version_id=version.id if version else None,
        row_count=validation["summary"]["total"],
        status="COMMITTED",
    )
    db.session.add(batch)

    created = updated = skipped = 0
    for row in validation["rows"]:
        if row["action"] == "ERROR":
            skipped += 1
            continue
        if row["action"] == "SKIP":
            skipped += 1
            continue
        outcome = committer(row["data"], options)
        if outcome == "CREATE":
            created += 1
        elif outcome == "UPDATE":
            updated += 1
        else:
            skipped += 1

    batch.created_count = created
    batch.updated_count = updated
    batch.skipped_count = skipped
    batch.error_count = validation["summary"]["error"]
    db.session.commit()
    return batch


def _source_doc_id(options):
    doc = options.get("source_document")
    return doc.id if doc else None


def _c_schedule(data, options):
    version = options["version"]
    is_baseline = bool(version.is_contractual_baseline)
    activity = WbsActivity.query.filter_by(
        schedule_version_id=version.id, wbs_code=data["wbs_code"]).first()
    outcome = "UPDATE" if activity else "CREATE"
    if activity is None:
        activity = WbsActivity(schedule_version_id=version.id, wbs_code=data["wbs_code"])
        db.session.add(activity)
        activity.sort_index = (options.get("_counter") or {}).get("n", 0)

    counter = options.setdefault("_counter", {"n": 0})
    counter["n"] += 1
    activity.sort_index = counter["n"]

    activity.parent_wbs_code = data["parent_wbs_code"] or None
    activity.activity_name = data["activity_name"]
    activity.work_package = data["work_package"] or None
    activity.discipline = data["discipline"] or None
    activity.level = data["wbs_code"].count(".") + 1

    # Contractual dates only ever live on a contractual baseline version.
    if is_baseline:
        activity.baseline_start = data["contractual_start"] or data["plan_start"]
        activity.baseline_finish = data["contractual_finish"] or data["plan_finish"]
        activity.plan_start = data["plan_start"] or activity.baseline_start
        activity.plan_finish = data["plan_finish"] or activity.baseline_finish
    else:
        activity.plan_start = data["plan_start"] or data["contractual_start"]
        activity.plan_finish = data["plan_finish"] or data["contractual_finish"]

    activity.duration_days = data["duration_days"]
    activity.is_milestone = data["is_milestone"]
    activity.unit = data["unit"] or None
    activity.total_required_quantity = data["total_required_quantity"]
    if data["progress_weight"] is not None:
        activity.progress_weight = data["progress_weight"]
        activity.weight_basis = "APPROVED WEIGHT"
    elif not activity.weight_basis:
        activity.weight_basis = "NOT SET"
    activity.progress_method = "QUANTITY" if data["total_required_quantity"] else "MANUAL"
    if data["reported_completion_pct"] is not None:
        activity.reported_completion_pct = data["reported_completion_pct"]
        activity.manual_pct = True
    activity.responsible_party = data["responsible_party"] or None
    activity.source_document_id = _source_doc_id(options) or version.source_document_id
    if data["source_reference"]:
        activity.notes = data["source_reference"]
    return outcome


def _c_areas(data, options):
    area = Area.query.filter_by(area_code=data["area_code"]).first()
    outcome = "UPDATE" if area else "CREATE"
    if area is None:
        area = Area(area_code=data["area_code"])
        db.session.add(area)
    area.area_name = data["area_name"]
    area.description = data["description"] or None
    area.drawing_reference = data["drawing_reference"] or None
    area.ifc_revision = data["ifc_revision"] or None
    area.active = data["active"]
    area.source_document_id = _source_doc_id(options)
    db.session.flush()
    if data["parent_area"]:
        parent = Area.query.filter_by(area_code=data["parent_area"]).first()
        area.parent_area_id = parent.id if parent else None
    return outcome


def _c_quantities(data, options):
    row = ActivityQuantity.query.filter_by(
        wbs_code=data["wbs_code"] or None, item=data["item"]).first()
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = ActivityQuantity(wbs_code=data["wbs_code"] or None, item=data["item"])
        db.session.add(row)
    row.area_id = data["area_id"]
    row.activity_name = data["activity_name"] or None
    row.total_quantity = data["total_quantity"] or 0.0
    row.unit = data["unit"] or None
    row.source_reference = data["source_reference"] or None
    row.revision = data["revision"] or None
    row.notes = data["notes"] or None
    row.source_document_id = _source_doc_id(options)
    return outcome


def _c_materials(data, options):
    package = ProcurementPackage.query.filter_by(package_code=data["package_code"]).first()
    outcome = "CREATE"
    if package is None:
        package = ProcurementPackage(package_code=data["package_code"],
                                     package_name=data["package_name"])
        db.session.add(package)
    else:
        outcome = "UPDATE"
    package.package_name = data["package_name"] or package.package_name
    if data["planned_delivery"]:
        package.planned_delivery = data["planned_delivery"]
    if data["actual_delivery"]:
        package.actual_delivery = data["actual_delivery"]
    if data["po_reference"]:
        package.po_reference = data["po_reference"]
    package.fat_required = package.fat_required or data["fat_required"]
    if package.fat_required and package.fat_status in (None, "NOT REQUIRED"):
        package.fat_status = "NOT STARTED"
    package.source_document_id = _source_doc_id(options)
    db.session.flush()

    if not data["item"]:
        return outcome

    material = Material.query.filter_by(package_id=package.id, item=data["item"]).first()
    outcome = "UPDATE" if material else "CREATE"
    if material is None:
        material = Material(package_id=package.id, item=data["item"])
        db.session.add(material)
    material.manufacturer = data["manufacturer"] or None
    material.vendor = data["vendor"] or None
    material.approved_vendor = data["approved_vendor"]
    material.unit = data["unit"] or None
    material.contract_quantity = data["contract_quantity"]
    material.total_required = data["total_required"] or 0.0
    material.ordered = data["ordered"] or 0.0
    material.planned_delivery = data["planned_delivery"]
    material.actual_delivery = data["actual_delivery"]
    material.po_reference = data["po_reference"] or None
    material.fat_required = data["fat_required"]
    if material.fat_required and material.fat_status in (None, "NOT REQUIRED"):
        material.fat_status = "NOT STARTED"
    material.source_document_id = _source_doc_id(options)
    return outcome


def _c_quality(data, options):
    row = InspectionRequirement.query.filter_by(
        itp_reference=data["itp_reference"],
        wbs_code=data["wbs_code"] or None,
        inspection_type=data["inspection_type"] or None).first()
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = InspectionRequirement(itp_reference=data["itp_reference"])
        db.session.add(row)
    row.wbs_code = data["wbs_code"] or None
    row.work_package = data["work_package"] or None
    row.inspection_type = data["inspection_type"] or None
    row.point_type = data["point_type"]
    row.required_evidence = data["required_evidence"] or None
    row.acceptance_criterion = data["acceptance_criterion"] or None
    row.applicable_specification = data["applicable_specification"] or None
    row.discipline = data["discipline"] or None
    row.source_document_id = _source_doc_id(options)
    row.active = True
    return outcome


def _c_acceptance(data, options):
    gate = AcceptanceGate.query.filter_by(gate_code=data["gate_code"]).first()
    if gate is None:
        return "SKIP"
    item = AcceptanceGateItem.query.filter_by(
        gate_id=gate.id, item_name=data["item_name"]).first()
    outcome = "UPDATE" if item else "CREATE"
    if item is None:
        item = AcceptanceGateItem(gate_id=gate.id, item_name=data["item_name"])
        db.session.add(item)
        item.sequence = (AcceptanceGateItem.query.filter_by(gate_id=gate.id).count() + 1)
    item.item_code = data["item_code"] or None
    item.description = data["description"] or None
    item.category = data["category"] or None
    item.responsible_party = data["responsible_party"] or None
    item.target_date = data["target_date"]
    item.contract_reference = data["contract_reference"] or None
    item.source_document_id = _source_doc_id(options)
    return outcome


def _c_permits(data, options):
    row = PermitItem.query.filter_by(item_name=data["item_name"]).first()
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = PermitItem(item_name=data["item_name"])
        db.session.add(row)
    for field in ("authority", "responsibility", "required_for", "document_reference",
                  "blocker_impact", "comments"):
        setattr(row, field, data[field] or None)
    row.required_by_date = data["required_by_date"]
    row.issued_date = data["issued_date"]
    row.expiry_date = data["expiry_date"]
    row.status = data["status"]
    row.source_document_id = _source_doc_id(options)
    return outcome


def _c_payments(data, options):
    row = (PaymentMilestone.query.filter_by(milestone_code=data["milestone_code"]).first()
           if data["milestone_code"]
           else PaymentMilestone.query.filter_by(description=data["description"]).first())
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = PaymentMilestone(description=data["description"])
        db.session.add(row)
    row.milestone_code = data["milestone_code"] or None
    row.description = data["description"]
    row.sequence = data["sequence"]
    row.percentage = data["percentage"]
    row.wbs_code = data["wbs_code"] or None
    row.gate_code = data["gate_code"]
    row.package_code = data["package_code"] or None
    row.planned_date = data["planned_date"]
    row.forecast_date = data["forecast_date"]
    row.status = data["status"]
    row.comments = data["comments"] or None
    row.source_document_id = _source_doc_id(options)
    return outcome


def _c_observations(data, options):
    row = SiteObservation.query.filter_by(
        entry_date=data["entry_date"], observation=data["observation"]).first()
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = SiteObservation(entry_date=data["entry_date"],
                              observation=data["observation"])
        db.session.add(row)
    row.area_id = data["area_id"]
    row.wbs_code = data["wbs_code"] or None
    row.category = data["category"]
    row.severity = data["severity"]
    row.action_required = data["action_required"] or None
    row.responsible_party = data["responsible_party"] or None
    row.target_date = data["target_date"]
    row.status = data["status"]
    if row.status == "CLOSED" and row.closed_date is None:
        row.closed_date = data["entry_date"]
    return outcome


def _c_documents(data, options):
    row = DocumentRegisterItem.query.filter_by(
        title=data["title"], gate_code=data["gate_code"],
        document_number=data["document_number"] or None).first()
    outcome = "UPDATE" if row else "CREATE"
    if row is None:
        row = DocumentRegisterItem(title=data["title"])
        db.session.add(row)
    row.document_number = data["document_number"] or None
    row.category = data["category"] or None
    row.discipline = data["discipline"] or None
    row.wbs_code = data["wbs_code"] or None
    row.revision = data["revision"] or None
    row.status = data["status"]
    row.required_date = data["required_date"]
    row.gate_code = data["gate_code"]
    row.mandatory = data["mandatory"]
    row.folder_path = data["folder_path"] or None
    row.remarks = data["remarks"] or None
    row.source_document_id = _source_doc_id(options)
    return outcome


_COMMITTERS = {
    "schedule": _c_schedule,
    "areas": _c_areas,
    "quantities": _c_quantities,
    "materials": _c_materials,
    "quality": _c_quality,
    "acceptance": _c_acceptance,
    "permits": _c_permits,
    "documents": _c_documents,
    "payments": _c_payments,
    "observations": _c_observations,
}


def write_templates(directory=None):
    """(Re)write the blank CSV import templates to disk."""
    directory = Path(directory or current_app.config["IMPORT_TEMPLATE_DIR"])
    directory.mkdir(parents=True, exist_ok=True)
    written = []
    for spec in SPECS.values():
        path = directory / f"template_{spec.key}.csv"
        path.write_text(spec.template_csv(), encoding="utf-8")
        written.append(path)
    return written


def available_prepared_files():
    """Structured extracts prepared from the registered Bassignana documents."""
    directory = Path(current_app.config["IMPORT_READY_DIR"])
    if not directory.exists():
        return []
    files = []
    for path in sorted(directory.glob("*.csv")):
        files.append({
            "name": path.name,
            "path": str(path),
            "size": path.stat().st_size,
            "key": _guess_import_key(path.name),
        })
    return files


def _guess_import_key(name):
    """Map a prepared file name to an import key.

    The leading segment of the file name wins, so a file such as
    `acceptance_gates_Schedule04A-04D.csv` is not mistaken for a schedule
    import because the words "Schedule 04A" appear later in the name.
    """
    lowered = name.lower()
    prefix = lowered.split("_", 1)[0]
    if prefix in SPECS:
        return prefix
    for key in SPECS:
        if key in lowered:
            return key
    if "wbs" in lowered or "programme" in lowered or "timeline" in lowered:
        return "schedule"
    if "boq" in lowered or "quantit" in lowered:
        return "quantities"
    if "gate" in lowered:
        return "acceptance"
    if "itp" in lowered:
        return "quality"
    if "vendor" in lowered or "package" in lowered:
        return "materials"
    return None


def register_prepared_file(path):
    """Copy a prepared file into the import workspace and return a token."""
    source = Path(path)
    token = uuid.uuid4().hex
    target = import_workspace() / f"{token}__{source.name}"
    target.write_bytes(source.read_bytes())
    return token, source.name
